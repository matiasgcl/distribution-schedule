# Distribution Schedule computation for linear release of supply.
# Requires: openpyxl, numpy
# Output: A spreadsheet with basic tokenomics and full distribution
# Bugs? probably! feel free to reach me: matiasgcl@protonmail.com
# Does my work helped you? tip it!
# Ethereum network: 0x030db3517414a9708cb3B12e3C71655D0D51d887

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, numbers
from openpyxl.utils import get_column_letter
import numpy as np

def inputPct(aux):
    #aux = str(input())
    if aux[len(aux)-1] == '%':
        aux = float(aux[:-1])
    else:
        aux = float(aux)
    if (aux < 0) or (aux>100):
        print('ERROR: Percentage cannot be higher than 100% or negative, exiting!')
        exit()
    return(aux)

def addMonth(date): # data MUST be YYYY-MM-DD
    year = int(date[:4])
    month = int(date[5:7])
    day = date[8:10]
    month += 1
    if int(month) < 10:
        month = '0'+str(month)
    if int(month) == 13:
        year += 1
        month = '01'
    if int(day)>28:
        day = '28'
        print('WARNING: Genesis day includes a day over 28, to avoid conflict with February day is set as 28. Remember, this is just a relative date, teams usually do not perform unlocks on perfect timing.')
    return(str(year)+'-'+str(month)+'-'+day)

print('In order to build the distribution and allocation tables, please choose a method to introduce the data:')
print('Type 1 to load data from a file or 2 for manual input:')
method = input()
if method == '2':
    print('Please introduce the following data:')
    print('Name of the token [string]:')
    token = input()
    print('Total supply [comma separator is allowed]:')
    aux = input()
    maxSupply = int(aux.replace(',',''))
    print('Genesis date [format YYYY-MM-DD, ex: 2022-05-14]:')
    aux = input()
    if((len(aux)!=10) or (aux[4]!='-') or (aux[7]!='-')):
        print('ERROR: date is not in the required format')
        exit()
    genesis_date = aux
    print('Number of categories for the allocation [int]:')
    numCategories = int(input())
elif method == '1':
    print('Please introduce the filename containing the token data:')
    dataFile = input()
    with open(dataFile) as f:
        data = f.read().splitlines()
    f.close()
    fixedCom1 = 7
    fixedCom2 = 22
    token = data[fixedCom1+1]
    maxSupply = int(data[fixedCom1+2].replace(',',''))
    aux = data[fixedCom1+3]
    if((len(aux)!=10) or (aux[4]!='-') or (aux[7]!='-')):
        print('ERROR: date is not in the required format')
        exit()
    genesis_date = aux
    numCategories = int(data[fixedCom1+4])
    expLength = fixedCom2+7*numCategories+1
    assert (len(data) == expLength), 'Incorrect number of lines (have you deleted some comment line or did you put an extra empty line at the end?)'

else: exit()

name = ['' for i in range(numCategories)]
percentage = [0 for i in range(numCategories)]
initLock = [0 for i in range(numCategories)]
initUnlock = [0 for i in range(numCategories)]
period = [0 for i in range(numCategories)]
lengthUnlock = [0 for i in range(numCategories)]
tokens = [0 for i in range(numCategories)]

for i in range(numCategories):
    if method == '2':
        j = i+1
        print('\nFor Category number '+str(j)+' (out of '+str(numCategories)+'), please introduce:')
        print('Name of the category [example: Public Sale]:')
        name[i] = str(input())

        print('Percentage of TOTAL SUPPLY for this category [example: 10.5% or 10.5, both are valid]:')
        aux = input()
        percentage[i] = inputPct(str(aux))
        tokens[i] = percentage[i]/100.0*maxSupply

        print('Percentage of tokens FOR THIS CATEGORY unlocked at genesis (i.e. immediatly unlocked, can be zero; if all tokens are unlocked at genesis, introduce 100):')
        aux = input()
        initUnlock[i] = inputPct(str(aux))
        if initUnlock[i] != 100:
            print('Initial lock period in months (can be zero) [example: 3.5]:')
            initLock[i] = float(input())

            print('Periodicity for unlocks in months (can be zero if all tokens are released after locking period) [example: 12 if tokens are released one time per year, 1 if tokens are released monthly]:')
            period[i] = float(input())

            print('Total unlock length in months NOT including the locking period (can be zero if all tokens are released after locking period) [example: 3*12 if tokens of this category are released on a 3 year horizon NOT including the initial locking period]:')
            aux = float(eval(input()))
            if (aux < period[i]):
                print('ERROR: Total length cannot be less than periodicity for unlocks, exiting!')
                break
            else:
                lengthUnlock[i] = aux

            if (lengthUnlock[i]%period[i]!=0):
                print('WARNING: The given total lenght is not a multiple of the period, you may arrive to full distribution before the expected time')

    if method == '1':
        name[i] = data[fixedCom2+7*i+1]
        percentage[i] = inputPct(data[fixedCom2+7*i+2])
        tokens[i] = percentage[i]/100.0*maxSupply
        initUnlock[i] = inputPct(data[fixedCom2+7*i+3])
        if initUnlock[i] != 100:
            initLock[i] = float(data[fixedCom2+7*i+4])
            period[i] = float(data[fixedCom2+7*i+5])
            aux = float(eval(data[fixedCom2+7*i+6]))
            if (aux < period[i]):
                print('ERROR: Total length cannot be less than periodicity for unlocks, exiting!')
                break
            else:
                lengthUnlock[i] = aux
            if (lengthUnlock[i]%period[i]!=0):
                print('WARNING: The given total lenght is not a multiple of the period, you may arrive to full distribution before the expected time')

# Once the inputs are done, lets check if numbers fit
if(abs(sum(percentage)-100)>0.05):
    print('ERROR: Sum of allocation percentages ('+str(percentage)+') does not sum 100% (error over 0.05%), exiting!')
    exit()
if(abs(sum(percentage)-100)>0):
    print('WARNING: Sum of allocation percentages is not exactly 100% (is precisely '+str(sum(percentage))+'%, this implies the total tokens distributed are not exactly '+str(maxSupply)+', but '+str(int(sum(tokens))))

# Now lets do some computations
fullUnlock = [0 for i in range(numCategories)]
genesisUnlock = [0 for i in range(numCategories)]
lockedTokens = [0 for i in range(numCategories)]
unlockFactor = [0 for i in range(numCategories)]
nUnlocks = [0 for i in range(numCategories)]
unlockAmount = [0 for i in range(numCategories)]

for i in range(numCategories):
    fullUnlock[i] = lengthUnlock[i] + initLock[i]

    if initUnlock[i] != 0:
        genesisUnlock[i] = initUnlock[i]/100*tokens[i]

    lockedTokens[i] = tokens[i] - genesisUnlock[i]

    if period[i]>0:
        nUnlocks[i] = int(lengthUnlock[i]/period[i] + 1)
        if (initUnlock[i] != 0) and (initLock[i] == 0):
            # if some portion of tokens (but not all, so period>0) are unlocked at genesis, then the first unlock is that portion and the rest is evenly distributed for the rest of unlocks.
            nUnlocks[i] -= 1
        unlockFactor[i] = 1./nUnlocks[i]
        unlockAmount[i] = unlockFactor[i]*lockedTokens[i]

    # the very special case when whitepaper is stated as 'no lock, coins released in some time with some period'
    if (initUnlock[i] == 0) and (initLock[i] == 0):
        genesisUnlock[i] = unlockAmount[i]
        lockedTokens[i] = tokens[i] - genesisUnlock[i]

totalMonths = int(max(fullUnlock))
unlocks = np.zeros([numCategories,totalMonths+1])

for i in range(numCategories):
    for j in range(totalMonths+1):
        if period[i]!=0:
            unlocks[i,j] = (j>0)*(j>=initLock[i])*(j<=fullUnlock[i])*((j-initLock[i])%period[i]==0)*unlockAmount[i]+(j==0)*genesisUnlock[i]
        else:
            unlocks[i,j] = (j==0)*genesisUnlock[i]

sums = [sum(unlocks[:,i]) for i in range(totalMonths+1)]

currentSupply = [0 for i in range(totalMonths+1)]
previousSupply = [0 for i in range(totalMonths+1)]
pctUnlockSupply = [0 for i in range(totalMonths+1)]
for i in range(totalMonths+1):
    if i==0:
        currentSupply[i] = sums[0]
        previousSupply[i] = 0
        pctUnlockSupply[i] = ''
    else:
        currentSupply[i] = currentSupply[i-1] + sums[i]
        previousSupply[i] = currentSupply[i-1]
        pctUnlockSupply[i] = sums[i]/previousSupply[i]

unlockDates = [0 for i in range(totalMonths+1)]
for i in range(totalMonths+1):
    if i==0: unlockDates[i] = genesis_date
    else: unlockDates[i] = addMonth(unlockDates[i-1])

# Create workbook
wb = Workbook()
# style for borders
marked=Side(border_style='thin',color="000000")
border=Border(top=marked,bottom=marked,left=marked,right=marked)

# add_sheet is used to create sheet.
sheet1 = wb.active
sheet1.title = 'Basic Data for '+token+' token'
sheet1.cell(row=1, column=1, value=token+' Tokenomics - Genesis Date: '+genesis_date+' - Total Supply: '+str(maxSupply))
sheet1.merge_cells('A1:H1')
sheet1['A1'].alignment = Alignment(horizontal="center", vertical="center",shrink_to_fit=True)
sheet1.cell(row=2, column=1, value='Category').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=2, value='Proportion').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=3, value='Total Tokens').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=4, value='Init Lock').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=5, value='Release Scheme').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=6, value='Genesis Unlock').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=7, value='Locked at Genesis').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=8, value='Unlock per Cat Period').alignment = Alignment(horizontal="center")
rang=sheet1['A1':'H2']
for cell in rang:
    for x in cell:
        x.border=border

for i in range(numCategories):

    sheet1.cell(row=i+3, column=1, value=name[i])
    sheet1.cell(row=i+3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet1.cell(row=i+3, column=2, value=float(percentage[i]/100.))
    sheet1.cell(row=i+3, column=2).number_format = numbers.FORMAT_PERCENTAGE_00
    sheet1.cell(row=i+3, column=3, value=int(tokens[i]))
    sheet1.cell(row=i+3, column=3).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    sheet1.cell(row=i+3, column=4, value=str(int(initLock[i]))+' Months')
    sheet1.cell(row=i+3, column=4).alignment = Alignment(horizontal="center", vertical="center")
    sheet1.cell(row=i+3, column=5, value='Each '+str(int(period[i]))+'Mo for '+str(int(lengthUnlock[i]))+'Mo')
    sheet1.cell(row=i+3, column=5).alignment = Alignment(horizontal="center", vertical="center")
    sheet1.cell(row=i+3, column=6, value=genesisUnlock[i])
    sheet1.cell(row=i+3, column=6).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    sheet1.cell(row=i+3, column=7, value=lockedTokens[i])
    sheet1.cell(row=i+3, column=7).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    sheet1.cell(row=i+3, column=8, value=unlockAmount[i])
    sheet1.cell(row=i+3, column=8).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    for j in range(8):
        sheet1.cell(row=i+3, column=j+1).border = border

sheet1.cell(row=numCategories+3, column=1, value='Sum').alignment = Alignment(horizontal="center")
sheet1.cell(row=numCategories+3, column=1).border = border
sheet1.cell(row=numCategories+3, column=2, value=float(sum(percentage)/100.))
sheet1.cell(row=numCategories+3, column=2).number_format = numbers.FORMAT_PERCENTAGE_00
sheet1.cell(row=numCategories+3, column=2).border = border
sheet1.cell(row=numCategories+3, column=3, value=sum(tokens))
sheet1.cell(row=numCategories+3, column=3).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
sheet1.cell(row=numCategories+3, column=3).border = border
sheet1.cell(row=numCategories+3, column=6, value=sum(genesisUnlock))
sheet1.cell(row=numCategories+3, column=6).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
sheet1.cell(row=numCategories+3, column=6).border = border
sheet1.cell(row=numCategories+3, column=7, value=sum(lockedTokens))
sheet1.cell(row=numCategories+3, column=7).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
sheet1.cell(row=numCategories+3, column=7).border = border

sheet1.column_dimensions['A'].width = 22
sheet1.column_dimensions['B'].width = 22
sheet1.column_dimensions['C'].width = 22
sheet1.column_dimensions['D'].width = 22
sheet1.column_dimensions['E'].width = 22
sheet1.column_dimensions['F'].width = 22
sheet1.column_dimensions['G'].width = 22
sheet1.column_dimensions['H'].width = 22

sheet2 = wb.create_sheet("Release Schedule", 0)
sheet2.title = 'Release Schedule for '+token+' token'
sheet2.cell(row=1, column=1, value=token+' Tokenomics - Genesis Date: '+genesis_date+' - Total Supply: '+str(maxSupply))
sheet2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7+numCategories)
sheet2['A1'].alignment = Alignment(horizontal="center", vertical="center",shrink_to_fit=True)
sheet2['A1'].border = border
sheet2.cell(row=2, column=1, value='Month').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=1).border = border
sheet2.column_dimensions['A'].width = 12
sheet2.cell(row=2, column=2, value='Approx unlock date').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=2).border = border
sheet2.column_dimensions['B'].width = 20
for i in range(numCategories):
    sheet2.cell(row=2, column=3+i, value=name[i]).alignment = Alignment(horizontal="center")
    sheet2.cell(row=2, column=3+i).border = border
    sheet2.column_dimensions[get_column_letter(3+i)].width = 18
sheet2.cell(row=2, column=3+numCategories, value='Total to be unlocked').alignment = Alignment(horizontal="center",shrink_to_fit=True)
sheet2.cell(row=2, column=3+numCategories).border = border
sheet2.column_dimensions[get_column_letter(3+numCategories)].width = 22
sheet2.cell(row=2, column=4+numCategories, value='Previous supply').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=4+numCategories).border = border
sheet2.column_dimensions[get_column_letter(4+numCategories)].width = 20
sheet2.cell(row=2, column=5+numCategories, value='% of previous supply').alignment = Alignment(horizontal="center",shrink_to_fit=True)
sheet2.cell(row=2, column=5+numCategories).border = border
sheet2.column_dimensions[get_column_letter(5+numCategories)].width = 22
sheet2.cell(row=2, column=6+numCategories, value='New supply').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=6+numCategories).border = border
sheet2.column_dimensions[get_column_letter(6+numCategories)].width = 20
sheet2.cell(row=2, column=7+numCategories, value='Circ % of total').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=7+numCategories).border = border
sheet2.column_dimensions[get_column_letter(7+numCategories)].width = 20

for i in range(totalMonths+1):
    sheet2.cell(row=3+i, column=1, value=i).alignment = Alignment(horizontal="center")
    sheet2.cell(row=3+i, column=1).border = border
    sheet2.cell(row=3+i, column=2, value=unlockDates[i]).alignment = Alignment(horizontal="center")
    sheet2.cell(row=3+i, column=2).border = border
    for j in range(numCategories):
        sheet2.cell(row=3+i, column=3+j, value=unlocks[j,i])
        sheet2.cell(row=3+i, column=3+j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet2.cell(row=3+i, column=4+j, value=sums[i])
        sheet2.cell(row=3+i, column=4+j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet2.cell(row=3+i, column=5+j, value=previousSupply[i])
        sheet2.cell(row=3+i, column=5+j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet2.cell(row=3+i, column=6+j, value=pctUnlockSupply[i])
        sheet2.cell(row=3+i, column=6+j).number_format = numbers.FORMAT_PERCENTAGE_00
        sheet2.cell(row=3+i, column=7+j, value=currentSupply[i])
        sheet2.cell(row=3+i, column=7+j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet2.cell(row=3+i, column=8+j, value=currentSupply[i]/maxSupply)
        sheet2.cell(row=3+i, column=8+j).number_format = numbers.FORMAT_PERCENTAGE_00
        for l in range(6):
            sheet2.cell(row=3+i,column=3+l+j).border = border
        sheet2.cell(row=4+totalMonths, column=3+j, value=sum(unlocks[j,:]))
        sheet2.cell(row=4+totalMonths, column=3+j).border = border
        sheet2.cell(row=4+totalMonths, column=3+j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet2.cell(row=4+totalMonths, column=4+j, value=sum(sums))
        sheet2.cell(row=4+totalMonths, column=4+j).border = border
        sheet2.cell(row=4+totalMonths, column=4+j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
sheet2.cell(row=4+totalMonths,column=1, value='Totals').alignment = Alignment(horizontal="center")
sheet2.cell(row=4+totalMonths,column=1).border = border

saveas = 'Release Schedule - '+token+'.xls'
wb.save(filename=saveas)
print('Success! \nData is saved in file: '+saveas)
